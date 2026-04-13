"""
Zemara — Maliyet & Gelir Analizi Excel Üreticisi
Çalıştır: python3 generate-excel.py
Fiyatlar: TR ₺150 Premium / ₺400 VIP | Intl $5.99 Premium / $12.99 VIP
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1
import os

# ── Renk sabitleri ───────────────────────────────────────────────────────────
C_HDR_DARK   = "1A0533"
C_HDR_MID    = "4A1070"
C_HDR_LIGHT  = "7B2FBE"
C_ACCENT     = "C9A0FF"
C_GREEN_DARK  = "166534"
C_GREEN_BG   = "DCFCE7"
C_RED_DARK   = "991B1B"
C_RED_BG     = "FEE2E2"
C_YELLOW_BG  = "FEF9C3"
C_YELLOW_DK  = "854D0E"
C_ROW_ALT    = "F3E8FF"
C_WHITE      = "FFFFFF"
C_GRAY       = "F9FAFB"
C_SECTION_BG = "EDE9FE"

KALI = {"top": "thin", "right": "thin", "bottom": "thin", "left": "thin"}

def wb_border():
    side = Side(style="thin", color="D1D5DB")
    return Border(left=side, right=side, top=side, bottom=side)

def sfill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def hdr(ws, row, col, value, bg=C_HDR_MID, fg=C_WHITE, bold=True, sz=11, wrap=False, center=True):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold, size=sz, color=fg)
    cell.fill = sfill(bg)
    cell.alignment = Alignment(horizontal="center" if center else "left",
                               vertical="center", wrap_text=wrap)
    cell.border = wb_border()
    return cell

def val(ws, row, col, value, bg=None, bold=False, sz=11, fmt=None,
        color=None, wrap=False, center=True):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold, size=sz, color=color or "000000")
    if bg:
        cell.fill = sfill(bg)
    cell.alignment = Alignment(horizontal="center" if center else "left",
                               vertical="center", wrap_text=wrap)
    cell.border = wb_border()
    if fmt:
        cell.number_format = fmt
    return cell

def section_hdr(ws, row, ncols, text, bg=C_SECTION_BG):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(bold=True, size=11, color=C_HDR_DARK)
    c.fill = sfill(bg)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = wb_border()

def title_row(ws, row, ncols, text, bg=C_HDR_DARK):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(bold=True, size=13, color=C_WHITE)
    c.fill = sfill(bg)
    c.alignment = Alignment(horizontal="center", vertical="center")

def subtitle_row(ws, row, ncols, text):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(italic=True, size=10, color="4B5563")
    c.fill = sfill("F5F3FF")
    c.alignment = Alignment(horizontal="left", vertical="center")

# ── Sabitler ─────────────────────────────────────────────────────────────────
FIXED  = 3164   # Aylık sabit altyapı gideri (₺)
KUR    = 38     # $1 = ₺38

# Komisyon sonrası + OpenAI sonrası net katkı PER PAID USER
NET = {
    "premTR":   137,   # ₺150 × (1-0.035) - ₺7.5 OpenAI
    "vipTR":    366,   # ₺400 × (1-0.035) - ₺20 OpenAI
    "premIntl": 186,   # $5.99 net after LS - OpenAI
    "vipIntl":  425,   # $12.99 net after LS - OpenAI
}

def scenario(pTR, vTR, pI, vI, marketing):
    gross    = pTR*150 + vTR*400 + pI*round(5.99*KUR) + vI*round(12.99*KUR)
    contrib  = pTR*NET["premTR"] + vTR*NET["vipTR"] + pI*NET["premIntl"] + vI*NET["vipIntl"]
    total_pd = pTR+vTR+pI+vI
    net      = contrib - FIXED - marketing
    status   = "✅ HEDEF" if net >= 70000 else ("🟡 YEŞİL" if net >= 50000 else "❌")
    return total_pd, round(gross), round(contrib), FIXED, marketing, round(net), status

wb = openpyxl.Workbook()
wb.remove(wb.active)

# ═════════════════════════════════════════════════════════════════════════════
# SHEET 1 — ÖZET DASHBOARD
# ═════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("📊 Özet Dashboard")
NC = 7

title_row(ws, 1, NC, "ZEMARA — MALİYET & GELİR ANALİZİ")
subtitle_row(ws, 2, NC, "Hazırlık tarihi: Nisan 2026  |  Hedef: ₺70.000 net/ay  |  Kur: $1 = ₺38")

# Fiyatlandırma
section_hdr(ws, 4, NC, "FİYATLANDIRMA")
headers_pr = ["", "TR Aylık (₺)", "TR Yıllık (₺)", "TR İndirim",
              "Intl Aylık ($)", "Intl Yıllık ($)", "Intl İndirim"]
for c, h in enumerate(headers_pr, 1):
    hdr(ws, 5, c, h, bg=C_HDR_LIGHT)

pricing_rows = [
    ("Ücretsiz", 0, 0, "—", 0, 0, "—"),
    ("Premium",  150, 1500, "%17", 5.99, 59.99, "%17"),
    ("VIP",      400, 4000, "%17", 12.99, 129.99, "%17"),
]
for i, row_data in enumerate(pricing_rows, 6):
    bg = C_ROW_ALT if i % 2 == 0 else C_WHITE
    for c, v in enumerate(row_data, 1):
        val(ws, i, c, v, bg=bg)

# Sabit giderler
section_hdr(ws, 10, NC, "SABİT AYLIK GİDERLER")
hdr_g = ["Kalem", "USD ($)", "TRY (₺)", "Not", "", "", ""]
for c, h in enumerate(hdr_g, 1):
    hdr(ws, 11, c, h, bg=C_HDR_LIGHT)
giders = [
    ("Vercel Pro", 20, 760, "Analitik + bandwidth + custom domains"),
    ("Supabase Pro", 25, 950, "Auth + DB + realtime + storage"),
    ("OpenAI API (~300 paid)", 25, 950, "≈ $0.083/kullanıcı/ay"),
    ("Domain / misc", 5, 190, "Yıllık $60 ÷ 12"),
    ("Apple Developer", 8.25, 314, "$99/yıl ÷ 12"),
    ("TOPLAM ALTYAPI", 83.25, 3164, "Sabit gider alt sınırı"),
]
for i, (kalem, usd, try_, not_) in enumerate(giders, 12):
    bg = C_RED_BG if "TOPLAM" in kalem else (C_ROW_ALT if i % 2 == 0 else C_WHITE)
    bold = "TOPLAM" in kalem
    val(ws, i, 1, kalem, bg=bg, bold=bold, center=False)
    val(ws, i, 2, usd, bg=bg, bold=bold, fmt='#,##0.00')
    val(ws, i, 3, try_, bg=bg, bold=bold, fmt='#,##0')
    val(ws, i, 4, not_, bg=bg, center=False)
    for c in range(5, 8):
        val(ws, i, c, "", bg=bg)

# Pazarlama bütçeleri
section_hdr(ws, 19, NC, "PAZARLAMA GİDERLERİ")
hdr(ws, 20, 1, "Dönem", bg=C_HDR_LIGHT)
hdr(ws, 20, 2, "Aylık Bütçe (₺)", bg=C_HDR_LIGHT)
hdr(ws, 20, 3, "Ne Kadar Sürer?", bg=C_HDR_LIGHT)
for c in range(4, 8):
    hdr(ws, 20, c, "", bg=C_HDR_LIGHT)
mkt_rows = [
    ("Ay 1–3 (büyüme fazı)", 12000, "3 ay"),
    ("Ay 4–6 (olgunlaşma)", 6000, "3 ay"),
    ("Ay 7+ (sürdürülebilir)", 3500, "Uzun vadeli"),
]
for i, (donem, butce, sure) in enumerate(mkt_rows, 21):
    bg = C_ROW_ALT if i % 2 == 0 else C_WHITE
    val(ws, i, 1, donem, bg=bg, center=False)
    val(ws, i, 2, butce, bg=bg, fmt='#,##0')
    val(ws, i, 3, sure, bg=bg)
    for c in range(4, 8):
        val(ws, i, c, "", bg=bg)

# Komisyonlar
section_hdr(ws, 25, NC, "ÖDEME KOMİSYONLARI")
hdr_k = ["Sağlayıcı", "Komisyon %", "Sabit/İşlem", "Kullanım", "", "", ""]
for c, h in enumerate(hdr_k, 1):
    hdr(ws, 26, c, h, bg=C_HDR_LIGHT)
komis = [
    ("iyzico (TR)",           "%3.3",  "₺0.25",  "TR kullanıcıları"),
    ("Lemon Squeezy (Intl)",  "%5.0",  "$0.50",  "Uluslararası kullanıcılar"),
    ("Apple IAP",             "%30",   "—",      "iOS içi satış VARSA zorunlu"),
    ("Google Play",           "%30",   "—",      "Android içi satış VARSA zorunlu"),
]
for i, (prov, comm, flat, note) in enumerate(komis, 27):
    bg = C_RED_BG if "Apple" in prov or "Google" in prov else (C_ROW_ALT if i % 2 == 0 else C_WHITE)
    val(ws, i, 1, prov, bg=bg, center=False)
    val(ws, i, 2, comm, bg=bg)
    val(ws, i, 3, flat, bg=bg)
    val(ws, i, 4, note, bg=bg, center=False)
    for c in range(5, 8):
        val(ws, i, c, "", bg=bg)

ws.column_dimensions["A"].width = 30
ws.column_dimensions["B"].width = 16
ws.column_dimensions["C"].width = 18
ws.column_dimensions["D"].width = 42
for col in ["E","F","G"]:
    ws.column_dimensions[col].width = 14
ws.row_dimensions[1].height = 28

# ═════════════════════════════════════════════════════════════════════════════
# SHEET 2 — ₺70K HEDEF SENARYOLAR
# ═════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("🎯 70K Hedef Senaryolar")
NC2 = 12
title_row(ws2, 1, NC2, "ZEMARA — ₺70.000 NET KÂR HEDEFİ: SENARYO ANALİZİ")
subtitle_row(ws2, 2, NC2,
    "Sabit: ₺3.164/ay | Net katkı: Prem TR ₺137 | VIP TR ₺366 | Prem Intl ₺186 | VIP Intl ₺425")

col_hdrs = ["SENARYO", "P.TR", "VIP TR", "P.Intl", "VIP Intl",
            "Toplam Ü.", "Brüt (₺)", "Net Katkı (₺)", "Altyapı (₺)",
            "Pazarlama (₺)", "NET KÂR (₺)", "DURUM"]
for c, h in enumerate(col_hdrs, 1):
    hdr(ws2, 3, c, h, bg=C_HDR_MID)

scenarios = [
    # (label, pTR, vTR, pIntl, vIntl, mkt)
    # ── Group 1: Steady state ──
    ("── STEADY STATE (Paz. ₺3.500) ──", None),
    ("%100 Premium TR",           500,   0,   0,   0, 3500),
    ("%100 VIP TR",                 0, 200,   0,   0, 3500),
    ("%70 Prem / %30 VIP TR",     270,  80,   0,   0, 3500),
    ("%50 Prem / %50 VIP TR",     170, 100,   0,   0, 3500),
    ("%30 Prem / %70 VIP TR",      90, 180,   0,   0, 3500),
    ("Karma: %50 TR / %50 Intl",  150,  75, 130,  30, 3500),
    ("Ağırlıklı VIP + Intl",       80,  60, 100,  80, 3500),
    ("Sadece VIP Intl",             0,   0,   0, 165, 3500),
    ("Optimal Mix (öneri) ⭐",     120,  80,  80,  50, 3500),
    # ── Group 2: Büyüme fazı ──
    ("── BÜYÜME FAZI Ay 1-3 (Paz. ₺12.000) ──", None),
    ("10 paid user (başlangıç)",    8,   2,   0,   0, 12000),
    ("30 paid user",               20,   8,   2,   0, 12000),
    ("60 paid user",               40,  15,   5,   0, 12000),
    ("100 paid user",              60,  25,  10,   5, 12000),
    # ── Group 3: Olgunlaşma ──
    ("── OLGUNLAŞMA Ay 4-6 (Paz. ₺6.000) ──", None),
    ("150 paid user",              90,  35,  15,  10, 6000),
    ("200 paid user",             110,  50,  25,  15, 6000),
    ("280 paid user",             150,  70,  40,  20, 6000),
    ("350 paid user",             180,  90,  55,  25, 6000),
    # ── Group 4: Kırılma noktaları ──
    ("── HEDEF KIRILMA NOKTALARI (Paz. ₺3.500) ──", None),
    ("Min. sadece VIP TR",          0, 202,   0,   0, 3500),
    ("Min. %50/%50 mix TR",       200, 110,   0,   0, 3500),
    ("Min. Optimal Mix ⭐",       120,  80,  80,  50, 3500),
    ("Min. Ağırlıklı Intl",        50,  30, 100,  80, 3500),
]

r = 4
for item in scenarios:
    if len(item) == 2:  # section header
        section_hdr(ws2, r, NC2, item[0])
        r += 1
        continue
    label, pTR, vTR, pI, vI, mkt = item
    total_pd, gross, contrib, fixed, mkt_v, net, status = scenario(pTR, vTR, pI, vI, mkt)
    row_bg = C_GREEN_BG if net >= 70000 else (C_YELLOW_BG if net >= 50000 else (C_RED_BG if net < 0 else C_WHITE))
    is_bold = net >= 70000
    data = [label, pTR, vTR, pI, vI, total_pd, gross, contrib, fixed, mkt_v, net, status]
    for c, v in enumerate(data, 1):
        fmt = "#,##0" if c in (7, 8, 9, 10, 11) else None
        clr = C_GREEN_DARK if net >= 70000 and c == 11 else (C_RED_DARK if net < 0 and c == 11 else None)
        val(ws2, r, c, v, bg=row_bg, bold=is_bold, fmt=fmt, color=clr,
            center=(c != 1))
    r += 1

ws2.column_dimensions["A"].width = 38
for col in ["B","C","D","E","F"]:
    ws2.column_dimensions[col].width = 9
for col in ["G","H","I","J","K"]:
    ws2.column_dimensions[col].width = 15
ws2.column_dimensions["L"].width = 12
ws2.row_dimensions[1].height = 28

# ═════════════════════════════════════════════════════════════════════════════
# SHEET 3 — 18 AYLIK PROJEKSİYON
# ═════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("📈 18 Ay Projeksiyon")
NC3 = 13
title_row(ws3, 1, NC3, "ZEMARA — 18 AYLIK GELİR & KÂR PROJEKSİYONU")
subtitle_row(ws3, 2, NC3,
    "Kur: $1=₺38 | Büyüme: Organik + Ücretli Pazarlama | Conversion free→paid: %2-3 hedef")

prjhdrs = ["AY", "Ücretsiz K.", "Paid Toplam", "Prem TR",
           "VIP TR", "Prem Intl", "VIP Intl",
           "Brüt Gelir (₺)", "Net Katkı (₺)", "Altyapı (₺)", "Pazarlama (₺)",
           "Aylık Net (₺)", "Kümülatif (₺)"]
for c, h in enumerate(prjhdrs, 1):
    hdr(ws3, 3, c, h, bg=C_HDR_MID)

months = [
    (1,   300,   5,  2,  1,  0, 12000),
    (2,   700,  12,  5,  2,  1, 12000),
    (3,  1400,  25, 10,  5,  2, 12000),
    (4,  2500,  50, 20, 10,  5,  6000),
    (5,  4000,  85, 32, 18,  8,  6000),
    (6,  6000, 130, 48, 28, 12,  6000),
    (7,  8500, 175, 65, 40, 18,  3500),
    (8, 11000, 215, 80, 50, 22,  3500),
    (9, 14000, 255, 95, 62, 28,  3500),
    (10,17000, 285,110, 72, 32,  3500),
    (11,20000, 310,122, 80, 36,  3500),
    (12,23000, 330,130, 85, 38,  3500),
    (13,26000, 345,138, 90, 42,  3500),
    (14,29000, 360,145, 95, 45,  3500),
    (15,32000, 370,150,100, 48,  3500),
    (16,35000, 378,155,103, 50,  3500),
    (17,38000, 384,158,106, 52,  3500),
    (18,41000, 388,162,108, 54,  3500),
]
cum = 0
for r_i, (ay, free, pTR, vTR, pI, vI, mkt) in enumerate(months, 4):
    gross   = pTR*150 + vTR*400 + pI*round(5.99*KUR) + vI*round(12.99*KUR)
    contrib = pTR*NET["premTR"] + vTR*NET["vipTR"] + pI*NET["premIntl"] + vI*NET["vipIntl"]
    net_m   = contrib - FIXED - mkt
    cum    += net_m
    tp      = pTR+vTR+pI+vI
    row_data = [ay, free, tp, pTR, vTR, pI, vI,
                round(gross), round(contrib), FIXED, mkt, round(net_m), round(cum)]
    bg = C_GREEN_BG if net_m >= 70000 else (C_YELLOW_BG if net_m >= 40000 else
         (C_RED_BG if net_m < 0 else (C_ROW_ALT if r_i % 2 == 0 else C_WHITE)))
    for c, v in enumerate(row_data, 1):
        fmt = "#,##0" if c >= 8 else ("#,##0" if c in (2, 3) else None)
        val(ws3, r_i, c, v, bg=bg, fmt=fmt)

ws3.column_dimensions["A"].width = 6
ws3.column_dimensions["B"].width = 14
ws3.column_dimensions["C"].width = 14
for col in ["D","E","F","G"]:
    ws3.column_dimensions[col].width = 10
for col in ["H","I","J","K","L","M"]:
    ws3.column_dimensions[col].width = 15
ws3.row_dimensions[1].height = 28

# ═════════════════════════════════════════════════════════════════════════════
# SHEET 4 — ÖDEME KOMİSYON ANALİZİ
# ═════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("💳 Ödeme Komisyonları")
NC4 = 7
title_row(ws4, 1, NC4, "ZEMARA — ÖDEME KOMİSYON ANALİZİ")
subtitle_row(ws4, 2, NC4, "Lemon Squeezy: %5 + $0.50/işlem  |  iyzico: %3.3 + ₺0.25/işlem")

# LS tablosu
section_hdr(ws4, 4, NC4, "LEMON SQUEEZY (Uluslararası) — %5 + $0.50/işlem")
ls_hdrs = ["Plan", "Fiyat ($)", "LS %5 ($)", "LS Sabit ($)", "Net Alınan ($)", "Eff. Kom %", "TRY Karşılığı (₺)"]
for c, h in enumerate(ls_hdrs, 1):
    hdr(ws4, 5, c, h, bg=C_HDR_LIGHT)

ls_plans = [
    ("Premium Monthly", 5.99),
    ("Premium Yearly",  59.99),
    ("VIP Monthly",     12.99),
    ("VIP Yearly",      129.99),
]
for i, (name, price) in enumerate(ls_plans, 6):
    comm = round(price * 0.05, 2)
    flat = 0.50
    net_ls = round(price - comm - flat, 2)
    eff_pct = f"%{(comm+flat)/price*100:.1f}"
    try_ = round(net_ls * KUR)
    bg = C_ROW_ALT if i % 2 == 0 else C_WHITE
    row_ls = [name, price, comm, flat, net_ls, eff_pct, try_]
    for c, v in enumerate(row_ls, 1):
        val(ws4, i, c, v, bg=bg, fmt="#,##0.00" if isinstance(v, float) else ("#,##0" if isinstance(v, int) else None))

# iyzico tablosu
section_hdr(ws4, 11, NC4, "İYZİCO (Türkiye) — %3.3 + ₺0.25/işlem")
iy_hdrs = ["Plan", "Fiyat (₺)", "iyzico %3.3 (₺)", "iyzico Sabit (₺)", "Net Alınan (₺)", "Eff. Kom %", "—"]
for c, h in enumerate(iy_hdrs, 1):
    hdr(ws4, 12, c, h, bg=C_HDR_LIGHT)

iy_plans = [
    ("Premium Monthly", 150),
    ("Premium Yearly",  1500),
    ("VIP Monthly",     400),
    ("VIP Yearly",      4000),
]
for i, (name, price) in enumerate(iy_plans, 13):
    comm = round(price * 0.033, 2)
    flat = 0.25
    net_iy = round(price - comm - flat, 2)
    eff_pct = f"%{(comm+flat)/price*100:.1f}"
    bg = C_ROW_ALT if i % 2 == 0 else C_WHITE
    row_iy = [name, price, comm, flat, net_iy, eff_pct, ""]
    for c, v in enumerate(row_iy, 1):
        val(ws4, i, c, v, bg=bg, fmt="#,##0.00" if isinstance(v, float) else ("#,##0" if isinstance(v, int) else None))

# Platform karşılaştırma
section_hdr(ws4, 18, NC4, "PLATFORM KOMİSYON KARŞILAŞTIRMASI")
plat_hdrs = ["Platform", "Komisyon %", "Sabit/İşlem", "Min. Fiyat", "Notlar", "", ""]
for c, h in enumerate(plat_hdrs, 1):
    hdr(ws4, 19, c, h, bg=C_HDR_LIGHT)
platforms = [
    ("iyzico",          "%3.3", "₺0.25",  "₺50+", "TR için ideal, etkili"),
    ("Lemon Squeezy",   "%5.0", "$0.50",  "$7+",  "$5.99 borderline, $12.99 iyi"),
    ("Stripe",          "%2.9", "$0.30",  "Her",  "TR kredi kartı sorunlu"),
    ("Paddle",          "%5.0", "$0.50",  "$8+",  "LS alternatifi"),
    ("Apple IAP",       "%30",  "—",      "Her",  "⚠️ iOS SAT = ZORUNLU"),
    ("Google Play IAP", "%30",  "—",      "Her",  "⚠️ Android SAT = ZORUNLU"),
]
for i, (pl, comm, flat, mn, note) in enumerate(platforms, 20):
    bg = C_RED_BG if "Apple" in pl or "Google" in pl else (C_GREEN_BG if "iyzico" in pl else (C_ROW_ALT if i%2==0 else C_WHITE))
    for c, v in enumerate([pl, comm, flat, mn, note, "", ""], 1):
        val(ws4, i, c, v, bg=bg, center=(c!=1 and c!=5))

section_hdr(ws4, 27, NC4,
    "⚠️ ÇÖZÜM: iOS/Android içi satış = %30 komisyon. Web (Vercel) üzerinden satış yap, app sadece içerik göstersin!")

ws4.column_dimensions["A"].width = 24
ws4.column_dimensions["B"].width = 14
ws4.column_dimensions["C"].width = 18
ws4.column_dimensions["D"].width = 16
ws4.column_dimensions["E"].width = 38
ws4.column_dimensions["F"].width = 12
ws4.column_dimensions["G"].width = 18
ws4.row_dimensions[1].height = 28

# ═════════════════════════════════════════════════════════════════════════════
# SHEET 5 — OPENAİ MALİYET ANALİZİ
# ═════════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("🤖 OpenAI Maliyetleri")
NC5 = 7
title_row(ws5, 1, NC5, "ZEMARA — OPENAİ API MALİYET PROJEKSİYONU (gpt-4o-mini)")
subtitle_row(ws5, 2, NC5,
    "Input: $0.15/1M token  |  Output: $0.60/1M token  |  Ort. istek: $0.000592  |  Kur: ₺38")

section_hdr(ws5, 4, NC5, "ÖZELLİK BAŞINA MALİYET")
feat_hdrs = ["Özellik", "Input Token", "Output Token", "Input ($)", "Output ($)", "Toplam/İstek ($)", "Aylık/User ($)"]
for c, h in enumerate(feat_hdrs, 1):
    hdr(ws5, 5, c, h, bg=C_HDR_LIGHT)

features = [
    ("Günlük Yorum",           600,  500, 0.000090, 0.000300, 0.000390),
    ("Haftalık Yorum",         750,  650, 0.000113, 0.000390, 0.000503),
    ("Natal Harita Yorumu",    950,  850, 0.000143, 0.000510, 0.000653),
    ("Transit Analizi",       1000,  900, 0.000150, 0.000540, 0.000690),
    ("Uyum Analizi",           950,  850, 0.000143, 0.000510, 0.000653),
    ("Tarot (3 kart)",          900,  800, 0.000135, 0.000480, 0.000615),
    ("Rüya Yorumu",             800,  700, 0.000120, 0.000420, 0.000540),
    ("Kristal Rehberi",         750,  650, 0.000113, 0.000390, 0.000503),
    ("ORTALAMA (karma)",        878,  767, 0.000132, 0.000460, 0.000592),
]
for i, (name, inp, out, in_cost, out_cost, total) in enumerate(features, 6):
    monthly_est = round(total * 3 * 30, 3)
    bg = C_SECTION_BG if "ORTALAMA" in name else (C_ROW_ALT if i%2==0 else C_WHITE)
    bold = "ORTALAMA" in name
    for c, v in enumerate([name, inp, out, in_cost, out_cost, total, monthly_est], 1):
        val(ws5, i, c, v, bg=bg, bold=bold,
            fmt=("0.000000" if c>=4 else None))

section_hdr(ws5, 16, NC5, "KULLANICI BAZLI AYLIK MALİYET PROJEKSİYONU")
user_hdrs = ["Paid K.", "Ücretsiz K.", "Günlük Aktif", "Aylık İstek", "Aylık ($)", "Aylık (₺)", "$/paid user"]
for c, h in enumerate(user_hdrs, 1):
    hdr(ws5, 17, c, h, bg=C_HDR_LIGHT)

user_rows = [
    (50, 500), (100, 1000), (200, 2000), (300, 3000),
    (500, 5000), (750, 8000), (1000, 12000),
]
for i, (paid, free) in enumerate(user_rows, 18):
    dau = round(paid*0.6 + free*0.15)
    mo_req = round(dau * 3 * 30)
    mo_cost_usd = round(mo_req * 0.000592, 2)
    mo_cost_try = round(mo_cost_usd * KUR)
    per_user = round(mo_cost_usd / paid, 3) if paid else 0
    bg = C_ROW_ALT if i%2==0 else C_WHITE
    for c, v in enumerate([paid, free, dau, mo_req, mo_cost_usd, mo_cost_try, per_user], 1):
        val(ws5, i, c, v, bg=bg,
            fmt=("#,##0" if c<=4 or c==6 else "0.000" if c==7 else "#,##0.00"))

for col, w in zip(["A","B","C","D","E","F","G"], [24,14,14,14,12,12,14]):
    ws5.column_dimensions[col].width = w
ws5.row_dimensions[1].height = 28

# ═════════════════════════════════════════════════════════════════════════════
# SHEET 6 — BREAK-EVEN & KÂR MARJI
# ═════════════════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("🔵 Break-Even Analizi")
NC6 = 11
title_row(ws6, 1, NC6, "ZEMARA — BREAK-EVEN & KÂR MARJI ANALİZİ")
subtitle_row(ws6, 2, NC6, "ARPU = Aylık Ortalama Gelir/Kullanıcı (mix bazlı)  |  Kür: ₺38=$1")

be_hdrs = ["Paid K.", "Mix", "ARPU (₺)", "Brüt (₺)", "Komisyon (₺)", "OpenAI (₺)",
           "Net Katkı (₺)", "Sabit+Mkt (₺)", "Net Kâr (₺)", "Kâr Marjı %", "MKT ROI %"]
for c, h in enumerate(be_hdrs, 1):
    hdr(ws6, 3, c, h, bg=C_HDR_MID)

def be_row(paid, mix_lbl, arpu, marketing):
    gross   = paid * arpu
    comm    = round(gross * 0.035)
    ai      = paid * 3
    contrib = round(gross - comm - ai)
    total_f = FIXED + marketing
    net_b   = contrib - total_f
    margin  = f"%{net_b/gross*100:.1f}" if gross > 0 else "—"
    roi     = f"%{net_b/marketing*100:.0f}" if marketing > 0 else "∞"
    return [paid, mix_lbl, arpu, round(gross), comm, ai, contrib, total_f, round(net_b), margin, roi]

groups = [
    ("── %100 Premium TR (ARPU ₺150) ──", 150, [30,60,100,150,200,280,350,450,550]),
    ("── %50 Premium / %50 VIP TR (ARPU ₺275) ──", 275, [20,40,80,120,180,240,300]),
    ("── %100 VIP TR (ARPU ₺400) ──", 400, [15,30,60,100,150,200,220]),
    ("── Optimal Mix TR+Intl (ARPU ₺280 tahmini) ──", 280, [40,80,130,180,250,300,340]),
]
row_n = 4
for group_lbl, arpu, user_list in groups:
    section_hdr(ws6, row_n, NC6, group_lbl)
    row_n += 1
    for n in user_list:
        mkt = 12000 if n < (80 if arpu>250 else 150) else (6000 if n < (180 if arpu>250 else 280) else 3500)
        data = be_row(n, f"ARPU ₺{arpu}", arpu, mkt)
        net_v = data[8]
        bg = C_GREEN_BG if net_v >= 70000 else (C_YELLOW_BG if net_v >= 40000 else
             (C_RED_BG if net_v < 0 else (C_ROW_ALT if row_n%2==0 else C_WHITE)))
        for c, v in enumerate(data, 1):
            fmt = "#,##0" if c in (3,4,5,6,7,8,9) else None
            clr = C_GREEN_DARK if net_v >= 70000 and c==9 else (C_RED_DARK if net_v < 0 and c==9 else None)
            val(ws6, row_n, c, v, bg=bg, fmt=fmt, color=clr, center=(c!=2))
        row_n += 1
    row_n += 1

for col, w in zip(["A","B","C","D","E","F","G","H","I","J","K"],
                  [9, 32, 11, 13, 14, 12, 14, 16, 13, 12, 12]):
    ws6.column_dimensions[col].width = w
ws6.row_dimensions[1].height = 28

# ═════════════════════════════════════════════════════════════════════════════
# SHEET 7 — MARKETİNG ROI
# ═════════════════════════════════════════════════════════════════════════════
ws7 = wb.create_sheet("📣 Marketing ROI")
NC7 = 10
title_row(ws7, 1, NC7, "ZEMARA — MARKETİNG ROI SENARYOLARI")
subtitle_row(ws7, 2, NC7,
    "Varsayım: %2.5 ücretsiz→ücretli dönüşüm  |  ARPU = ₺200 karma  |  CPM ₺8 Instagram/TikTok")

section_hdr(ws7, 4, NC7, "A — ÜCRETLI REKLAM (Instagram / TikTok)")
ad_hdrs = ["Bütçe (₺)", "CPM (₺)", "CTR %", "Landing CVR %", "İzlenme", "Tıklama", "Ücretsiz K.", "Paid K.", "Gelir (₺)", "ROI %"]
for c, h in enumerate(ad_hdrs, 1):
    hdr(ws7, 5, c, h, bg=C_HDR_LIGHT)

for i, budget in enumerate([3000, 6000, 10000, 15000, 20000], 6):
    cpm, ctr, cvr, arpu = 8, 2.5, 25, 200
    imp  = round((budget/cpm)*1000)
    clks = round(imp*ctr/100)
    free = round(clks*cvr/100)
    paid = round(free*0.025)
    rev  = paid*arpu
    roi  = f"%{(rev-budget)/budget*100:.0f}"
    bg   = C_GREEN_BG if rev > budget else (C_YELLOW_BG if rev > budget*0.5 else C_RED_BG)
    for c, v in enumerate([budget, cpm, f"%{ctr}", f"%{cvr}", imp, clks, free, paid, rev, roi], 1):
        val(ws7, i, c, v, bg=bg, fmt="#,##0" if c in (1,5,6,7,8,9) else None)

section_hdr(ws7, 12, NC7, "B — ASO (App Store Optimizasyonu) — Organik, Sıfır Maliyet")
aso_hdrs = ["Ay", "Organik İndirme", "Aktif K.", "Dönüşüm %", "Paid K.", "Gelir (₺)", "", "", "", ""]
for c, h in enumerate(aso_hdrs, 1):
    hdr(ws7, 13, c, h, bg=C_HDR_LIGHT)
aso_rows = [(1,200,150,"%2",4,800), (3,800,600,"%2.5",15,3000),
            (6,2500,1800,"%3",54,10800), (12,8000,5500,"%3",165,33000)]
for i, row in enumerate(aso_rows, 14):
    bg = C_ROW_ALT if i%2==0 else C_WHITE
    for c, v in enumerate(list(row)+["","","",""], 1):
        val(ws7, i, c, v, bg=bg, fmt="#,##0" if isinstance(v,int) else None)

section_hdr(ws7, 19, NC7, "C — İÇERİK MARKETİNG (TikTok/Reels Organik)")
cont_hdrs = ["Takipçi", "Aylık Erişim", "Siteye Gelen", "Kayıt %", "Ücretsiz K.", "Paid K.", "Gelir (₺)", "", "", ""]
for c, h in enumerate(cont_hdrs, 1):
    hdr(ws7, 20, c, h, bg=C_HDR_LIGHT)
cont_rows = [
    (1000,5000,100,"%30",30,1,200),
    (5000,30000,750,"%30",225,6,1200),
    (20000,150000,4500,"%30",1350,34,6800),
    (50000,400000,12000,"%30",3600,90,18000),
]
for i, row in enumerate(cont_rows, 21):
    bg = C_ROW_ALT if i%2==0 else C_WHITE
    for c, v in enumerate(list(row)+["","",""], 1):
        val(ws7, i, c, v, bg=bg, fmt="#,##0" if isinstance(v,int) else None)

section_hdr(ws7, 26, NC7, "D — ÖNERİLEN KANAL MİXİ (Ay 1-3)")
mix_hdrs = ["Kanal", "Aylık Bütçe (₺)", "Tahmini Paid K.", "Not", "", "", "", "", "", ""]
for c, h in enumerate(mix_hdrs, 1):
    hdr(ws7, 27, c, h, bg=C_HDR_LIGHT)
mix_rows = [
    ("TikTok/Instagram reklam", 6000, "15-20", "En hızlı dönüş"),
    ("ASO optimizasyonu", 0, "5-10", "Yatırımsız, ama zaman ister"),
    ("Organik içerik üretimi", 1000, "5-8", "Araç/ekipman maliyeti"),
    ("Influencer partnership", 3000, "10-15", "1-2 mikro influencer"),
    ("Google Ads (astrology)", 2000, "8-12", "Anahtar kelime hedefleme"),
    ("TOPLAM", 12000, "43-65", "İlk 3 ay tahmini"),
]
for i, (kanal, butce, paid, note) in enumerate(mix_rows, 28):
    bg = C_SECTION_BG if "TOPLAM" in kanal else (C_ROW_ALT if i%2==0 else C_WHITE)
    bold = "TOPLAM" in kanal
    for c, v in enumerate([kanal, butce, paid, note]+[""]*6, 1):
        val(ws7, i, c, v, bg=bg, bold=bold,
            fmt="#,##0" if c==2 and isinstance(butce,int) else None,
            center=(c!=1 and c!=4))

for col, w in zip(["A","B","C","D","E","F","G","H","I","J"],
                  [28, 16, 16, 38, 14, 14, 14, 12, 12, 10]):
    ws7.column_dimensions[col].width = w
ws7.row_dimensions[1].height = 28

# ═════════════════════════════════════════════════════════════════════════════
# SHEET 8 — RAKİP KARŞILAŞTIRMA
# ═════════════════════════════════════════════════════════════════════════════
ws8 = wb.create_sheet("🏆 Rakip Karşılaştırma")
NC8 = 7
title_row(ws8, 1, NC8, "ZEMARA — RAKİP FİYAT & ÖZELLİK KARŞILAŞTIRMASI")
subtitle_row(ws8, 2, NC8, "Kaynak: App Store / Resmi websiteler  |  Nisan 2026")

comp_hdrs = ["Uygulama", "Aylık ($)", "Yıllık ($)", "Model", "Ücretsiz Tier", "Harita?", "Platform"]
for c, h in enumerate(comp_hdrs, 1):
    hdr(ws8, 3, c, h, bg=C_HDR_MID)

comps = [
    ("CHANI",            12.00,  108,    "Hub (içerik+ses+ritüel)", "Temel chart, podcast", "❌",          "iOS/Android"),
    ("The Pattern",      14.99,  120,    "Hub (içerik kütüphanesi)", "Bazı içerikler",       "❌",          "iOS"),
    ("TimePassages",      7.99,   60,    "Chart tool",               "1 chart",              "❌ (web yok)", "iOS/Android/Web"),
    ("Astro Future",      4.99,   30,    "Chart tool",               "Temel chart",          "❌",          "iOS/Android"),
    ("AstroMatrix",       None,   14.99, "Chart + ads",             "Neredeyse her şey",    "❌",          "iOS/Android"),
    ("astro.com",         1.07,   12.90, "Web tool",                 "Tam ACG harita!",      "✅ ÜCRETSİZ", "Web"),
    ("Astro Gold iOS",    7.99,   40,    "Pro chart",               "Temel chart",          "✅ +$24.99",  "iOS"),
    ("Co-Star",           None,   None,  "Sosyal",                  "Her şey ücretsiz",     "❌",          "iOS/Android"),
    ("Nebula",            9.99,   50,    "Hub",                      "Temel",                "❌",          "iOS/Android"),
    ("Kasamba",           None,   None,  "Canlı okuma",              "3 dk ücretsiz",        "❌",          "iOS/Android/Web"),
]
for i, row in enumerate(comps, 4):
    bg = C_ROW_ALT if i%2==0 else C_WHITE
    for c, v in enumerate(row, 1):
        fmt = "#,##0.00" if isinstance(v, float) else None
        val(ws8, i, c, v or "—", bg=bg, fmt=fmt, center=(c!=1 and c!=4 and c!=5 and c!=6))

# Zemara rows
section_hdr(ws8, 15, NC8, "ZEMARA (bizim konumumuz)")
zemara_rows = [
    ("Zemara Premium (TR)",  "~$3.95",  "~$39.47", "Hub + Map + AI (10+ özellik)", "10 şehir + 3 AI/gün", "✅ DAHİL",  "iOS/Android/Web"),
    ("Zemara VIP (TR)",      "~$10.52", "~$105",   "Hub + Map + AI VIP",           "10 şehir + 3 AI/gün", "✅ + Derin","iOS/Android/Web"),
    ("Zemara Premium (Intl)","$5.99",   "$59.99",  "Hub + Map + AI (10+ özellik)", "10 şehir + 3 AI/gün", "✅ DAHİL",  "iOS/Android/Web"),
    ("Zemara VIP (Intl)",    "$12.99",  "$129.99", "Hub + Map + AI VIP",           "10 şehir + 3 AI/gün", "✅ + Derin","iOS/Android/Web"),
]
for i, row in enumerate(zemara_rows, 16):
    bg = C_GREEN_BG
    for c, v in enumerate(row, 1):
        val(ws8, i, c, v, bg=bg, bold=True, center=(c!=1 and c!=4 and c!=5))

section_hdr(ws8, 21, NC8,
    "SONUÇ: Zemara = CHANI ($12) / Pattern ($14.99) altında ama çok daha geniş kapsam. Intl $5.99 rekabetçi konumda.")

for col, w in zip(["A","B","C","D","E","F","G"],
                  [24, 12, 12, 28, 30, 14, 20]):
    ws8.column_dimensions[col].width = w
ws8.row_dimensions[1].height = 28

# ═════════════════════════════════════════════════════════════════════════════
# SHEET 9 — ÖZET YÖNERGELER
# ═════════════════════════════════════════════════════════════════════════════
ws9 = wb.create_sheet("✅ Özet & Yönergeler")
NC9 = 4
title_row(ws9, 1, NC9, "ZEMARA — ₺70K HEDEF: ÖZET YÖNERGELER")

section_hdr(ws9, 3, NC9, "KRİTİK SORULAR & CEVAPLAR")
qa_hdrs = ["Soru", "Cevap", "Not", ""]
for c, h in enumerate(qa_hdrs, 1):
    hdr(ws9, 4, c, h, bg=C_HDR_LIGHT)
qa_rows = [
    ("Kaç paid user (sadece VIP TR)?",     "~202 VIP",           "En az user = senaryo"),
    ("Kaç paid user (karma %50/%50)?",     "~315 kullanıcı",     "Daha stabil gelir"),
    ("Kaç paid user (sadece Premium)?",   "~555 kullanıcı",     "Zor conversion, yüksek hacim"),
    ("Intl dahil optimum?",                "~260 (karma mix)",   "TR+Intl dengeli"),
    ("Aylık pazarlama bütçesi (başta)?",  "₺12.000 (Ay 1-3)",  "Sonra ₺3.500'e iner"),
    ("Kaçıncı ayda ₺70K hedefi?",         "Ay 12-13",           "Gerçekçi senaryo"),
    ("Toplam ilk yıl yatırım?",            "~-₺80.000",          "Pazarlama kümülatif kayıp"),
]
for i, (soru, cevap, not_) in enumerate(qa_rows, 5):
    bg = C_ROW_ALT if i%2==0 else C_WHITE
    val(ws9, i, 1, soru, bg=bg, center=False)
    val(ws9, i, 2, cevap, bg=bg, bold=True)
    val(ws9, i, 3, not_, bg=bg, center=False)
    val(ws9, i, 4, "", bg=bg)

section_hdr(ws9, 13, NC9, "EN KRİTİK 5 FAKTÖR")
crit_hdrs = ["#", "Faktör", "Açıklama", ""]
for c, h in enumerate(crit_hdrs, 1):
    hdr(ws9, 14, c, h, bg=C_HDR_LIGHT)
crits = [
    (1, "VIP dönüşüm oranı", "%30 VIP mix → 220 user yeter. VIP özelliklerini güçlendir."),
    (2, "Web-first satış", "iOS/Android içi satış = %30 komisyon. Vercel web üzerinden sat!"),
    (3, "Intl kullanıcı", "CHANI $12, Pattern $14.99 varken $5.99 ile rakipleriz."),
    (4, "Churn kontrolü", "Aylık %5-10 churn normal. Her ay ~15-25 yeni paid user lazım."),
    (5, "LS verification", "Bekleme süresince Türkiye iyzico çalışıyor, intl gecikiyor."),
]
for i, (n, fak, acik) in enumerate(crits, 15):
    bg = C_ROW_ALT if i%2==0 else C_WHITE
    val(ws9, i, 1, n, bg=bg)
    val(ws9, i, 2, fak, bg=bg, bold=True)
    val(ws9, i, 3, acik, bg=bg, center=False, wrap=True)
    val(ws9, i, 4, "", bg=bg)

section_hdr(ws9, 21, NC9, "RİSK MATRİSİ")
risk_hdrs = ["Risk", "Olasılık", "Etki", "Önlem"]
for c, h in enumerate(risk_hdrs, 1):
    hdr(ws9, 22, c, h, bg=C_HDR_LIGHT)
risks = [
    ("LS verification gecikmesi", "Orta", "Yüksek", "Stripe Atlas araştır"),
    ("₺150 dönüşüm düşük", "Yüksek", "Yüksek", "A/B test: ₺99 vs ₺150 ilk ay"),
    ("OpenAI API fiyat artışı", "Düşük", "Orta", "Önbellekleme + rate limiting"),
    ("Rakip ücretsiz özellik artışı", "Orta", "Orta", "VIP exclusive özelliklere odaklan"),
    ("App Store IAP zorunluluğu", "Orta", "Çok Yüksek", "Web-first, app = içerik görüntüleyici"),
    ("Yüksek churn (>%15/ay)", "Orta", "Yüksek", "Onboarding + bildirim + retention içeriği"),
]
for i, (risk, olas, etki, onem) in enumerate(risks, 23):
    etki_bg = C_RED_BG if "Yüksek" in etki or "Çok" in etki else C_YELLOW_BG
    olas_bg = C_YELLOW_BG if "Orta" in olas else (C_RED_BG if "Yüksek" in olas else C_GREEN_BG)
    val(ws9, i, 1, risk, bg=C_ROW_ALT if i%2==0 else C_WHITE, center=False)
    val(ws9, i, 2, olas, bg=olas_bg)
    val(ws9, i, 3, etki, bg=etki_bg)
    val(ws9, i, 4, onem, bg=C_ROW_ALT if i%2==0 else C_WHITE, center=False)

ws9.row_dimensions[29].height = 20
for col, w in zip(["A","B","C","D"], [36, 14, 18, 44]):
    ws9.column_dimensions[col].width = w
ws9.row_dimensions[1].height = 28
for r in range(15, 30):
    ws9.row_dimensions[r].height = 30

# ─── KAYDET ──────────────────────────────────────────────────────────────────
out = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Zemara_Maliyet_Analizi.xlsx")
wb.save(out)
print(f"✅ Excel oluşturuldu: {out}")
print("📊 9 sheet:")
print("   1. Özet Dashboard      6. Break-Even Analizi")
print("   2. 70K Hedef Senaryolar 7. Marketing ROI")
print("   3. 18 Ay Projeksiyon    8. Rakip Karşılaştırma")
print("   4. Ödeme Komisyonları   9. Özet & Yönergeler")
print("   5. OpenAI Maliyetleri")
